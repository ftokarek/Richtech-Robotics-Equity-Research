

import openpyxl
import pandas as pd
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any
import re
from datetime import datetime


def get_filing_metadata(file_path: str) -> Dict[str, str]:
    
    filename = Path(file_path).stem
    
    metadata = {
        'filename': Path(file_path).name,
        'filepath': file_path
    }
    
    
    date_match = re.search(r'(\d{4}-\d{2}-\d{2})', filename)
    if date_match:
        metadata['filing_date'] = date_match.group(1)
    
    
    form_match = re.search(r'\(([^)]+)\)\s+\d{4}-\d{2}-\d{2}', filename)
    if form_match:
        metadata['form_code'] = form_match.group(1)
    
    
    ticker_match = re.search(r'^([A-Z]+)\s+\(([^)]+)\)', filename)
    if ticker_match:
        metadata['ticker'] = ticker_match.group(1)
        metadata['company'] = ticker_match.group(2)
    
    return metadata


def read_excel_with_merged_cells(file_path: str, sheet_name: str, 
                                  fill_merged: bool = True) -> pd.DataFrame:
    
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]
    
    
    data = []
    for row in ws.iter_rows(values_only=False):
        row_data = []
        for cell in row:
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        
                        master_cell = ws.cell(merged_range.min_row, merged_range.min_col)
                        row_data.append(master_cell.value if fill_merged else None)
                        break
            else:
                row_data.append(cell.value)
        data.append(row_data)
    
    wb.close()
    
    
    if data:
        df = pd.DataFrame(data)
    else:
        df = pd.DataFrame()
    
    return df


def find_sheets_by_keyword(file_path: str, keywords: List[str], 
                           case_sensitive: bool = False) -> List[str]:
    
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    sheet_names = wb.sheetnames
    wb.close()
    
    matching_sheets = []
    for sheet_name in sheet_names:
        for keyword in keywords:
            search_name = sheet_name if case_sensitive else sheet_name.lower()
            search_keyword = keyword if case_sensitive else keyword.lower()
            
            if search_keyword in search_name:
                matching_sheets.append(sheet_name)
                break
    
    return matching_sheets


def detect_table_boundaries(df: pd.DataFrame, 
                           min_non_null: int = 2) -> Tuple[int, int, int, int]:
    
    if df.empty:
        return (0, 0, 0, 0)
    
    
    row_has_data = df.notna().sum(axis=1) >= min_non_null
    rows_with_data = row_has_data[row_has_data].index.tolist()
    
    if not rows_with_data:
        return (0, 0, 0, 0)
    
    start_row = min(rows_with_data)
    end_row = max(rows_with_data)
    
    
    col_has_data = df.notna().sum(axis=0) >= min_non_null
    cols_with_data = col_has_data[col_has_data].index.tolist()
    
    if not cols_with_data:
        return (start_row, end_row, 0, 0)
    
    start_col = min(cols_with_data)
    end_col = max(cols_with_data)
    
    return (start_row, end_row, start_col, end_col)


def extract_table_from_sheet(file_path: str, sheet_name: str,
                             header_rows: int = 1,
                             skip_empty_rows: bool = True) -> pd.DataFrame:
    
    
    df = read_excel_with_merged_cells(file_path, sheet_name)
    
    if df.empty:
        return df
    
    
    start_row, end_row, start_col, end_col = detect_table_boundaries(df)
    
    if start_row >= end_row or start_col >= end_col:
        return pd.DataFrame()
    
    
    df = df.iloc[start_row:end_row+1, start_col:end_col+1]
    
    
    if header_rows > 0 and len(df) > header_rows:
        if header_rows == 1:
            df.columns = df.iloc[0]
            df = df.iloc[1:].reset_index(drop=True)
        else:
            
            header = df.iloc[:header_rows]
            
            new_cols = []
            for col in df.columns:
                col_headers = [str(h) if pd.notna(h) else '' for h in header[col]]
                combined = ' '.join([h for h in col_headers if h]).strip()
                new_cols.append(combined if combined else f'Column_{col}')
            df.columns = new_cols
            df = df.iloc[header_rows:].reset_index(drop=True)
    
    
    if skip_empty_rows:
        df = df.dropna(how='all').reset_index(drop=True)
    
    return df


def get_all_sheets_info(file_path: str) -> List[Dict[str, Any]]:
    
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    
    sheets_info = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        info = {
            'sheet_name': sheet_name,
            'dimensions': ws.dimensions,
            'max_row': ws.max_row,
            'max_column': ws.max_column
        }
        sheets_info.append(info)
    
    wb.close()
    return sheets_info


def extract_multi_level_headers(df: pd.DataFrame, 
                                header_rows: int = 2) -> pd.DataFrame:
    
    if len(df) < header_rows:
        return df
    
    header_data = df.iloc[:header_rows]
    new_columns = []
    
    for col in df.columns:
        parts = []
        for i in range(header_rows):
            val = header_data.iloc[i][col]
            if pd.notna(val) and str(val).strip():
                parts.append(str(val).strip())
        
        combined = ' - '.join(parts) if parts else f'Column_{col}'
        new_columns.append(combined)
    
    df.columns = new_columns
    df = df.iloc[header_rows:].reset_index(drop=True)
    
    return df


def safe_read_excel(file_path: str, sheet_name: str, **kwargs) -> Optional[pd.DataFrame]:
    
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, **kwargs)
        return df
    except Exception as e:
        print(f"Error reading sheet '{sheet_name}' from {file_path}: {e}")
        return None


def get_sheet_names(file_path: str) -> List[str]:
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    except Exception as e:
        print(f"Error getting sheet names from {file_path}: {e}")
        return []

